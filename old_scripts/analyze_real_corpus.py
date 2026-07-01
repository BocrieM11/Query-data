#!/usr/bin/env python3
"""
A: 提取関西方言テキスト → 对比分析
B: 全コーパスのフィラー/相槌/ポーズ頻度統計
C: 全Excel → ShareGPT JSONL変換
"""

import os
import json
import re
import glob
from pathlib import Path
from collections import Counter, defaultdict
import openpyxl

DATA_ROOT = Path("数据")
OUTPUT_DIR = Path("corpus_analysis")
OUTPUT_DIR.mkdir(exist_ok=True)

# ============================================================
# 工具関数
# ============================================================

def find_all_excels(data_root):
    """再帰的に全xlsxファイルを列挙（一時ファイル除外）"""
    excels = []
    for root, dirs, files in os.walk(data_root):
        for f in files:
            if f.endswith('.xlsx') and not f.startswith('~$'):
                excels.append(os.path.join(root, f))
    return sorted(excels)


def extract_conversations_from_xlsx(filepath):
    """1つのExcelファイルから会話テキストを抽出"""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        return []

    conversations = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 4:
            continue

        # カラムG（発話内容）とカラムF（話者）を検出
        text_col = None
        speaker_col = None
        for col_idx, cell in enumerate(rows[2] if len(rows) > 2 else []):
            val = str(cell).strip() if cell else ""
            if "発話" in val and "内容" in val:
                text_col = col_idx
            if "話者" in val or (col_idx > 0 and "話" in val and "者" in val):
                speaker_col = col_idx

        # ヒューリスティック: 最も長いテキスト列を発話列に
        if text_col is None:
            avg_lengths = {}
            for col in range(min(15, len(rows[0]) if rows else 0)):
                texts = []
                for row in rows[3:]:
                    if col < len(row) and row[col]:
                        texts.append(str(row[col]))
                if texts:
                    avg_lengths[col] = sum(len(t) for t in texts) / len(texts)
            if avg_lengths:
                text_col = max(avg_lengths, key=avg_lengths.get)

        # 話者列のヒューリスティック
        if speaker_col is None:
            for col in range(min(8, len(rows[0]) if rows else 0)):
                vals = set()
                for row in rows[3:]:
                    if col < len(row) and row[col]:
                        v = str(row[col]).strip()
                        if len(v) <= 6:
                            vals.add(v)
                if 2 <= len(vals) <= 10:
                    speaker_col = col
                    break

        if text_col is None:
            continue

        # 会話を抽出
        utterances = []
        for row in rows[3:]:
            if text_col >= len(row) or not row[text_col]:
                continue
            text = str(row[text_col]).strip()
            if not text or len(text) < 2:
                continue

            speaker = "unknown"
            if speaker_col is not None and speaker_col < len(row) and row[speaker_col]:
                speaker = str(row[speaker_col]).strip()

            utterances.append({
                "speaker": speaker,
                "text": text,
            })

        if utterances:
            conversations.append({
                "sheet": sheet_name,
                "utterances": utterances,
                "total_turns": len(utterances),
            })

    return conversations


def classify_directory(dirname):
    """ディレクトリ名から会話タイプを推定"""
    name = dirname

    # 判定可能なキーワード
    if "同性友人同士の謝罪" in name:
        return "apology_roleplay_FF"
    if "関西方言" in name or "嬨廈曽尵" in name:
        return "kansai_dialect"
    if "日本人と学習者" in name or "妛廗幰" in name:
        return "native_learner_mixed"
    if "仕事" in name or "悩み" in name or "怘帠" in name:
        return "work_worry_theme"

    # パターンマッチング
    patterns = {
        "01": "friends_MM_FF",
        "02": "firstmeeting_friends_FF",
        "06": "friends_FF",
        "14": "firstmeeting_hierarchical_FF_MM_audio",
        "29": "work_worry_theme_FF_MF_audio_video",
        "30": "native_learner_mixed_audio_video",
        "31": "kansai_dialect_firstmeeting_audio_video",
    }
    for prefix, label in patterns.items():
        if dirname.startswith(prefix):
            return label
    return "unknown"


# ============================================================
# A: 関西方言分析
# ============================================================

KANSAI_FEATURES = [
    # 文法
    (r'へん\b', '否定形 (〜ない→〜へん)'),
    (r'やで\b', '断定＋伝達 (〜だよ→〜やで)'),
    (r'やろ\b', '推量 (〜だろう→〜やろ)'),
    (r'どす\b', '丁寧断定 (京都: です→どす)'),
    (r'さかい\b', '理由 (〜から→〜さかい)'),
    (r'おる\b', '居る (いる→おる)'),
    (r'のう\b', '詠嘆 (ねぇ→のう)'),
    (r'わ\b', '終助詞 (女性)'),
    (r'まへん\b', '丁寧否定 (〜ません→〜まへん)'),
    (r'どした', 'どうした→どした'),
    (r'ほんま', '本当→ほんま'),
    (r'めっちゃ', 'とても→めっちゃ'),
    (r'あかん', '駄目→あかん'),
    (r'ちゃう', '違う→ちゃう'),
    (r'せや', 'そうだ→せや'),
    (r'おおきに', 'ありがとう→おおきに'),
]

FILLERS = [
    'あのー', 'あの', 'えーと', 'ええと', 'まあ', 'そのー', 'その',
    'なんか', 'なんやろ', 'なあ', 'ねえ', 'うーん', 'えー',
]

AIZUCHI = ['うん', 'はい', 'ええ', 'ああ', 'そう', 'へえ', 'ふーん', 'なるほど']


def analyze_kansai(all_utterances_by_dir):
    """関西方言ディレクトリのテキストを分析"""
    print("=" * 60)
    print("A: 関西方言 分析")
    print("=" * 60)

    kansai_utterances = []
    for dirname, convs in all_utterances_by_dir.items():
        if classify_directory(dirname) == "kansai_dialect":
            for conv in convs:
                for utt in conv["utterances"]:
                    kansai_utterances.append(utt["text"])

    if not kansai_utterances:
        # 他のディレクトリからも探す
        for dirname, convs in all_utterances_by_dir.items():
            for conv in convs:
                for utt in conv["utterances"]:
                    text = utt["text"]
                    if any(re.search(p, text)
                           for p in [r'へん\b', r'やで', r'ほんま', r'あかん', r'どす', r'さかい']):
                        kansai_utterances.append(text)

    print(f"\n関西方言 発話数: {len(kansai_utterances)}")

    if not kansai_utterances:
        print("  ⚠️ 関西方言の発話が見つかりませんでした")
        return {}

    # 特徴カウント
    feature_counts = {}
    for pattern, desc in KANSAI_FEATURES:
        count = sum(1 for t in kansai_utterances if re.search(pattern, t))
        if count > 0:
            feature_counts[desc] = {"count": count, "pct": count / len(kansai_utterances) * 100}

    print("\n関西方言特徴 出現頻度:")
    for desc, data in sorted(feature_counts.items(), key=lambda x: -x[1]["count"]):
        print(f"  {desc:30s} : {data['count']:4d}回 ({data['pct']:5.1f}%)")

    # サンプル
    print("\n関西方言サンプル:")
    for i, t in enumerate(kansai_utterances[:15]):
        print(f"  [{i+1}] {t[:100]}")

    return feature_counts


# ============================================================
# B: フィラー・相槌・ポーズ統計
# ============================================================

def analyze_linguistic_features(all_utterances_by_dir):
    """全コーパスの言語特徴を統計分析"""
    print("\n" + "=" * 60)
    print("B: フィラー・相槌・ポーズ 統計")
    print("=" * 60)

    all_texts = []
    all_turns = 0
    by_dir = defaultdict(lambda: {"turns": 0, "texts": [], "filler_count": 0, "aizuchi_count": 0})

    for dirname, convs in all_utterances_by_dir.items():
        dtype = classify_directory(dirname)
        for conv in convs:
            for utt in conv["utterances"]:
                text = utt["text"]
                all_texts.append(text)
                all_turns += 1
                by_dir[dtype]["turns"] += 1
                by_dir[dtype]["texts"].append(text)

    # フィラー統計
    print("\n--- フィラー出現頻度 (全コーパス) ---")
    filler_total = Counter()
    for text in all_texts:
        for f in FILLERS:
            if f in text:
                filler_total[f] += 1

    for filler, count in filler_total.most_common():
        pct = count / all_turns * 100
        bar = "█" * int(pct)
        print(f"  {filler:12s} : {count:4d}回 ({pct:5.1f}%) {bar}")

    # 相槌統計
    print("\n--- 相槌 出現頻度 (全コーパス) ---")
    aizuchi_total = Counter()
    for text in all_texts:
        for a in AIZUCHI:
            # 完全一致（単独発話としての相槌）
            if text.strip() == a or text.strip().startswith(a + '。') or text.strip().startswith(a + '、'):
                aizuchi_total[a] += 1

    for aizuchi, count in aizuchi_total.most_common():
        pct = count / all_turns * 100
        bar = "█" * int(pct / 2)
        print(f"  {aizuchi:12s} : {count:4d}回 ({pct:5.1f}%) {bar}")

    # ポーズ・笑い・言い直し
    print("\n--- パラ言語マーカー ---")
    markers = {
        "ポーズ(,,)": sum(1 for t in all_texts if ',,' in t),
        "笑い<笑い>": sum(1 for t in all_texts if '<笑い>' in t),
        "笑いながら": sum(1 for t in all_texts if '笑いながら' in t),
        "重複開始{<}": sum(1 for t in all_texts if '{<' in t),
        "言い直し(同一語反復)": sum(1 for t in all_texts if re.search(r'(\S{2,})、?\1', t)),
    }
    for marker, count in markers.items():
        pct = count / all_turns * 100
        print(f"  {marker:25s} : {count:4d}回 ({pct:5.1f}%)")

    # 発話長分布
    lengths = [len(t) for t in all_texts if len(t) > 1]
    avg_len = sum(lengths) / len(lengths) if lengths else 0
    print(f"\n--- 発話統計 ---")
    print(f"  総発話数: {all_turns}")
    print(f"  平均発話長: {avg_len:.1f}文字")
    print(f"  最短: {min(lengths)}字, 最長: {max(lengths)}字")

    # あなたの対話との比較用データを生成
    print(f"\n--- 推奨フィラー密度（あなたの対話の調整用） ---")
    total_filler_rate = sum(filler_total.values()) / all_turns * 100
    print(f"  実際のフィラー密度: {total_filler_rate:.1f}% (発話あたり)")
    print(f"  実際の相槌密度: {sum(aizuchi_total.values()) / all_turns * 100:.1f}%")

    return {
        "filler_counts": dict(filler_total.most_common()),
        "aizuchi_counts": dict(aizuchi_total.most_common()),
        "total_turns": all_turns,
        "avg_length": avg_len,
        "filler_density": total_filler_rate,
    }


# ============================================================
# C: 全Excel → ShareGPT JSONL
# ============================================================

def convert_all_to_sharegpt(all_utterances_by_dir, filepath_map):
    """全データをShareGPT形式に変換"""
    print("\n" + "=" * 60)
    print("C: Excel → ShareGPT JSONL 変換")
    print("=" * 60)

    all_records = []
    by_type = defaultdict(list)
    stats = defaultdict(lambda: {"conversations": 0, "utterances": 0})

    for dirname, convs in all_utterances_by_dir.items():
        dtype = classify_directory(dirname)

        for conv in convs:
            # 話者を human/gpt に割り当て
            speakers = set(u["speaker"] for u in conv["utterances"])
            speaker_map = {}
            for i, sp in enumerate(sorted(speakers)):
                speaker_map[sp] = f"speaker_{i+1}"

            conversations = []
            for utt in conv["utterances"]:
                conversations.append({
                    "from": "human",
                    "value": utt["text"],
                })

            if not conversations:
                continue

            record = {
                "conversations": conversations,
                "source_type": dtype,
                "source_dir": dirname[:60],
                "source_file": os.path.basename(
                    filepath_map.get(id(conv), "unknown.xlsx")
                ),
                "sheet": conv.get("sheet", ""),
                "total_turns": conv["total_turns"],
                "speakers": sorted(speakers),
                "dataset": "Japanese_Natural_Conversation_Corpus",
            }

            all_records.append(record)
            by_type[dtype].append(record)
            stats[dtype]["conversations"] += 1
            stats[dtype]["utterances"] += conv["total_turns"]

    # 保存：全量
    all_path = OUTPUT_DIR / "real_corpus_all.jsonl"
    with open(all_path, "w", encoding="utf-8") as f:
        for r in all_records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    print(f"\n  ✅ all.jsonl: {len(all_records)} 会話 → {all_path}")

    # 保存：タイプ別
    type_dir = OUTPUT_DIR / "by_type"
    type_dir.mkdir(exist_ok=True)
    for dtype, recs in sorted(by_type.items()):
        fpath = type_dir / f"{dtype}.jsonl"
        with open(fpath, "w", encoding="utf-8") as f:
            for r in recs:
                f.write(json.dumps(r, ensure_ascii=False) + "\n")
        print(f"  ✅ {dtype}: {len(recs)} 会話 → {fpath}")

    # 統計
    print(f"\n--- タイプ別集計 ---")
    for dtype, s in sorted(stats.items()):
        print(f"  {dtype:40s} : {s['conversations']:3d}会話, {s['utterances']:4d}発話")

    return all_records


# ============================================================
# Main
# ============================================================

def main():
    print("=" * 60)
    print("日本語自然会話コーパス 分析 + 変換パイプライン")
    print("=" * 60)

    # 全Excelファイルを探す
    all_excels = find_all_excels(DATA_ROOT)
    print(f"\n📂 {len(all_excels)} Excelファイルを検出")

    # 全データを抽出
    all_utterances_by_dir = defaultdict(list)
    filepath_map = {}

    for i, fpath in enumerate(all_excels):
        dirname = os.path.basename(os.path.dirname(os.path.dirname(fpath)))
        # 中間ディレクトリ構造を保持
        rel_path = os.path.relpath(fpath, DATA_ROOT)
        convs = extract_conversations_from_xlsx(fpath)
        for conv in convs:
            filepath_map[id(conv)] = fpath
        all_utterances_by_dir[dirname].extend(convs)

        if (i + 1) % 20 == 0:
            print(f"  ... {i+1}/{len(all_excels)} ファイル処理済み")

    total_convs = sum(len(v) for v in all_utterances_by_dir.values())
    total_utts = sum(
        sum(c["total_turns"] for c in convs)
        for convs in all_utterances_by_dir.values()
    )
    print(f"  抽出完了: {total_convs} 会話, {total_utts} 発話")

    # A: 関西方言
    kansai_features = analyze_kansai(all_utterances_by_dir)

    # B: フィラー統計
    ling_stats = analyze_linguistic_features(all_utterances_by_dir)

    # C: ShareGPT変換
    sharegpt_records = convert_all_to_sharegpt(all_utterances_by_dir, filepath_map)

    # サマリー
    print(f"\n{'=' * 60}")
    print(f"✅ 全タスク完了")
    print(f"{'=' * 60}")
    print(f"\n出力:")
    print(f"  {OUTPUT_DIR / 'real_corpus_all.jsonl'}  ← 全量ShareGPT")
    print(f"  {OUTPUT_DIR / 'by_type/'}           ← タイプ別ShareGPT")
    print(f"\n次のステップ:")
    print(f"  1. 統計結果を元に対話のフィラー密度を調整")
    print(f"  2. 関西方言サンプルを画像別対話の方言表現に反映")
    print(f"  3. real_corpus_all.jsonl を補助訓練データとして使用")


if __name__ == "__main__":
    main()
